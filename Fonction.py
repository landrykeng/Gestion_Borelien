import streamlit as st
import pandas as pd
from datetime import datetime, date, time
#import time
import os
import uuid
from pathlib import Path
import openpyxl
from openpyxl import Workbook
import base64
import gspread
from google.oauth2.service_account import Credentials
from functools import lru_cache
import json

st.session_state.data_loaded = False
version="local"  # "local" or "online"
SERVICE_ACCOUNT_FILE = st.secrets["gcp_service_account"] if version=="online" else  "Credential.json"
SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive"
]
SHEET_NAME = "Gestion_Borelien"  # Nom de votre Google Sheet

# Variables globales pour la mise en cache
_client = None
_sheet = None
_last_connection = 0
CONNECTION_TIMEOUT = 300  # 5 minutes
statut_df = pd.DataFrame({"ID": ["A+","A","A-","N+","N","N-"], "Honnoraire": [5000, 4500, 4000, 3500, 3000, 2500]})

@lru_cache(maxsize=1)
def get_credentials():
    """Cache les credentials pour éviter de les recharger"""
    try:
        # SERVICE_ACCOUNT_FILE est ici un dictionnaire contenant les credentials
        if version=="online":
            return Credentials.from_service_account_info(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
        else:
            return Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    except Exception as e:
        st.error(f"Erreur credentials : {e}")
        return None

def get_google_sheets_client():
    """Client Google Sheets avec cache et réutilisation de connexion"""
    global _client, _last_connection
    
    current_time=int(datetime.now().timestamp())
    
    # Réutiliser le client existant s'il est encore valide
    if _client and (current_time - _last_connection) < CONNECTION_TIMEOUT:
        return _client
    
    try:
        creds = get_credentials()
        if not creds:
            return None
            
        _client = gspread.authorize(creds)
        _last_connection = current_time
        return _client
        
    except Exception as e:
        st.error(f"Erreur connexion : {e}")
        _client = None
        return None

def get_google_sheet():
    """Sheet Google Sheets avec cache"""
    global _sheet
    
    if _sheet:
        return _sheet
    
    try:
        client = get_google_sheets_client()
        if not client:
            return None
            
        _sheet = client.open(SHEET_NAME)
        return _sheet
        
    except gspread.SpreadsheetNotFound:
        st.error(f"Google Sheet '{SHEET_NAME}' introuvable")
        return None
    except Exception as e:
        st.error(f"Erreur ouverture sheet : {e}")
        return None

def get_worksheet(sheet_name):
    """Récupère un onglet avec gestion d'erreur optimisée"""
    try:
        sheet = get_google_sheet()
        if not sheet:
            return None
        return sheet.worksheet(sheet_name)
    except gspread.WorksheetNotFound:
        st.warning(f"Onglet '{sheet_name}' introuvable")
        return None
    except Exception as e:
        st.error(f"Erreur onglet '{sheet_name}' : {e}")
        return None

def batch_init_google_sheet():
    """Initialise tous les onglets en une seule fois (plus rapide)"""
    try:
        sheet = get_google_sheet()
        if not sheet:
            return False
        
        existing_worksheets = {ws.title: ws for ws in sheet.worksheets()}
        
        # Configuration des onglets à créer
        sheets_config = {
            "Séances": ["ID", "Date", "Matière", "HeureArrivée", "HeureDepart", 
                       "Classe", "IntituléCours", "Centre", "idEnseignant", "NombreEtudiants"],
            "Étudiants": ["Matricule", "Nom", "Prénom", "Sexe", "Concours1", "Concours2", 
                         "Concours3", "Téléphone", "DateArrivée", "Etablissement", "Centre"]
        }
        
        # Créer tous les onglets manquants en une fois
        for sheet_name, headers in sheets_config.items():
            if sheet_name not in existing_worksheets:
                worksheet = sheet.add_worksheet(title=sheet_name, rows="1000", cols="20")
                worksheet.append_row(headers)
                st.success(f"Onglet '{sheet_name}' créé")
        
        return True
        
    except Exception as e:
        st.error(f"Erreur initialisation : {e}")
        return False

def save_to_google_sheet(sheet_name, data):
    """Version optimisée pour sauvegarder"""
    try:
        worksheet = get_worksheet(sheet_name)
        if not worksheet:
            # Tentative de création si onglet inexistant
            if batch_init_google_sheet():
                worksheet = get_worksheet(sheet_name)
            if not worksheet:
                return False
        
        # Optimisation selon le type de données
        if isinstance(data, list):
            worksheet.append_row(data)
        elif isinstance(data, pd.DataFrame) and not data.empty:
            # Utiliser batch update pour les gros volumes
            values = data.values.tolist()
            if len(values) > 10:  # Batch pour plus de 10 lignes
                worksheet.append_rows(values, value_input_option='USER_ENTERED')
            else:
                for row in values:
                    worksheet.append_row(row)
        else:
            st.warning("Données vides ou format non supporté")
            return False
        
        return True
        
    except Exception as e:
        st.error(f"Erreur sauvegarde '{sheet_name}' : {e}")
        return False

def read_from_google_sheet(sheet_name, use_cache=True):
    """Version optimisée pour lire avec cache optionnel"""
    try:
        worksheet = get_worksheet(sheet_name)
        if not worksheet:
            return pd.DataFrame()
        
        # Récupération optimisée
        if use_cache:
            # Utiliser get_all_values() qui est plus rapide que get_all_records()
            values = worksheet.get_all_values()
            if not values or len(values) < 2:
                return pd.DataFrame()
            
            # Créer DataFrame manuellement (plus rapide)
            headers = values[0]
            data = values[1:]
            df = pd.DataFrame(data, columns=headers)
            
            # Nettoyer les lignes vides
            df = df.replace('', pd.NA).dropna(how='all')
        else:
            # Méthode classique pour les cas où on veut les types automatiques
            data = worksheet.get_all_records()
            df = pd.DataFrame(data) if data else pd.DataFrame()
        
        return df
        
    except Exception as e:
        st.error(f"Erreur lecture '{sheet_name}' : {e}")
        return pd.DataFrame()


EXCEL_FILE = "stato_sphere_data.xlsx"

def get_base64_image(image_path):
        try:
            with open(image_path, "rb") as img_file:
                return base64.b64encode(img_file.read()).decode()
        except FileNotFoundError:
            st.error(f"Image non trouvée: {image_path}")
            return None

def read_from_excel(sheet_name):
    """Lit les données d'une feuille Excel"""
    try:
        if os.path.exists(EXCEL_FILE):
            df = pd.read_excel(EXCEL_FILE, sheet_name=sheet_name)
            return df
        else:
            return pd.DataFrame()
    except Exception as e:
        return pd.DataFrame()

def add_bg_local(image_file):
        with open(image_file, "rb") as image:
            encoded = base64.b64encode(image.read()).decode()
        st.markdown(
            f"""
            <style>
            .stApp {{
            background: linear-gradient(
                rgba(255,255,255,0.85), 
                rgba(255,255,255,0.85)
            ), url("data:image/png;base64,{encoded}");
            background-attachment: fixed;
            background-size: cover;
            }}
            </style>
            """,
            unsafe_allow_html=True
        )

# Chargement des bases
#@st.cache_data()
def load_all_data():
    """
    Charge toutes les données depuis Google Sheets et les stocke dans st.session_state.
    """
    if "data_loaded" not in st.session_state or not st.session_state.data_loaded:
        with st.spinner("Chargement des données...",show_time=True):
            st.session_state.etudiants_df = read_from_google_sheet("Étudiants")
            st.session_state.enseignants_df = read_from_google_sheet("Enseignants")
            st.session_state.seances_df = read_from_google_sheet("Séances")
            st.session_state.depenses_df = read_from_google_sheet("Dépenses")
            st.session_state.versements_df = read_from_google_sheet("Versements")
            st.session_state.ventes_df = read_from_google_sheet("Ventes_Bords")
            st.session_state.presence_df = read_from_google_sheet("Présences")
            st.session_state.presences_df = read_from_google_sheet("Présences")
            st.session_state.fiches_paie_df = read_from_google_sheet("Fiches_Paie")
            st.session_state.Connect_df = read_from_google_sheet("Connexion")
            st.session_state.data_loaded = True
            st.session_state.nb_open_sessions = 1
        
SHEETS = {
        "etudiant": "Étudiants",
        "depense": "Dépenses", 
        "versement": "Versements",
        "bord": "Ventes_Bords",
        "presence": "Présences",
        "enseignant": "Enseignants",
        "seance": "Séances",
        "statut": "Statuts"
    }
FILIERE_CHOICES = ["MATH L1", "MATH L2", "MATH L3", "PHY L1", "PHY L2", "PHY L3", "INFO L1", "INFO L2", "INFO L3"]
CENTRES_CHOICES = ["Douala", "Yaoundé", "Dschang"]
SEXE_CHOICES = ["Homme", "Femme"]
COURS_CHOICES = ["Algèbre linéaire 1", "Algèbre bilinéaire et multilinéaire", "Calcul différentiel 1", "Calcul différentiel 2", "Calcul différentiel 3",
                 "Calcul intégral 1","Calcul intégral 2", "Statistiques et probabilités", "Statistiques 2", "Physique", "Cryptographie","Variables complexes", "Equations différentielles"]
TYPE_DEPENSE_CHOICES = [
        "Publicité", "Dépense de lancement", "Impressions des fiches", 
        "Impressions des bords", "Motivations des associés", "Loyer", 
        "Accessoires", "Salaire"
    ]

STATUT_ENSEIGNANT_CHOICES=["A+", "A", "A-","N","N-","N--"]
ETABLISSEMENT=["Université Yde1","Université Yde2(SOA)","Université Dschang",
                "Université Dla","Polytech Yde","Polytech Dla","Ecole Normale","Autre"]
MOYEN_PAIEMENT_CHOICES=["Espèces","Mobile Money","Orange Money"]

dict_honnoraire = {"A+":5000,
                   "A":4500,
                   "A-":4000,
                   "N+":3500,
                   "N":3500,
                   "N-":2500}
    



def init_excel_file():
    """Initialise le fichier Excel avec les feuilles nécessaires"""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        wb.remove(wb.active)
        
        headers = {
            "Étudiants": ["Matricule", "Nom", "Prénom", "Sexe", "Concours1", "Concours2", "Concours3", 
                         "Téléphone", "DateArrivée", "Etablissement", "Centre"],
            "Séances": ["ID", "Date", "Matière", "HeureArrivée", "HeureDepart", "Classe", "IntituléCours", "Centre", "idEnseignant", "NombreEtudiants"],
            "Enseignants": ["ID", "Nom", "Prénom", "Matière", "Téléphone", "Statut", "TauxHoraire", "Centre", "DateEmbauche"]
        }
        
        for sheet_name, header in headers.items():
            ws = wb.create_sheet(title=sheet_name)
            ws.append(header)
        
        wb.save(EXCEL_FILE)

def save_to_excel(sheet_name, data):
    """Sauvegarde les données dans la feuille Excel spécifiée"""
    try:
        init_excel_file()
        wb = openpyxl.load_workbook(EXCEL_FILE)
        
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(title=sheet_name)
            if sheet_name == "Séances":
                ws.append(["ID", "Date", "Matière", "HeureArrivée", "HeureDepart", "Classe", "IntituléCours", "Centre", "idEnseignant", "NombreEtudiants"])
            elif sheet_name == "Étudiants":
                ws.append(["Matricule", "Nom", "Prénom", "Sexe", "Concours1", "Concours2", "Concours3", "Téléphone", "DateArrivée", "Etablissement", "Centre"])
        else:
            ws = wb[sheet_name]
        
        ws.append(data)
        wb.save(EXCEL_FILE)
        return True
    except Exception as e:
        st.error(f"Erreur lors de la sauvegarde : {e}")
        return False

def generate_matricule():
    """Génère un matricule unique pour un étudiant"""
    return f"STAT{datetime.now().strftime('%Y%m%d')}{str(uuid.uuid4())[:4].upper()}"

def get_teacher_info(username):
    """Récupère les informations de l'enseignant connecté"""
    teachers_info = {
        "mbang.pierre": {
            "nom": "Dr. MBANG Pierre",
            "matieres": ["MATHEMATIQUES", "STATISTIQUES"],
            "centres": ["Yaoundé", "Dschang"],
            "statut": "Permanent"
        },
        "njoya.marie": {
            "nom": "Prof. NJOYA Marie",
            "matieres": ["FRANCAIS", "CULTURE GENERALE"],
            "centres": ["Douala", "Yaoundé"],
            "statut": "Permanent"
        },
        "fomba.jean": {
            "nom": "M. FOMBA Jean",
            "matieres": ["ECONOMIE", "COMPTABILITE"],
            "centres": ["Dschang"],
            "statut": "Vacataire"
        },
        "admin": {
            "nom": "Enseignant Test",
            "matieres": COURS_CHOICES,
            "centres": CENTRES_CHOICES,
            "statut": "Test"
        }
    }
    
    return teachers_info.get(username, {
        "nom": f" {username}",
        "matieres": COURS_CHOICES,
        "centres": CENTRES_CHOICES,
        "statut": "Non défini"
    })


def search_student(query, etudiants_df):
    """Recherche un étudiant par nom ou prénom"""
    if etudiants_df.empty or not query:
        return pd.DataFrame()
    
    query = query.lower().strip()
    
    # Recherche dans nom et prénom
    mask = (
        etudiants_df['Nom'].str.lower().str.contains(query, na=False) |
        etudiants_df['Prénom'].str.lower().str.contains(query, na=False) |
        etudiants_df['Matricule'].str.lower().str.contains(query, na=False)
    )
    
    return etudiants_df[mask]

def calculate_student_stats(matricule, versements_df, presences_df, seances_df):
    """Calcule les statistiques d'un étudiant"""
    stats = {
        "total_verse": 0,
        "nombre_versements": 0,
        "taux_absenteisme": 0,
        "total_seances": 0,
        "seances_presentes": 0,
        "seances_absentes": 0,
        "seances_retard": 0
    }
    
    # Calcul des versements
    if not versements_df.empty and 'idMatricule' in versements_df.columns:
        student_versements = versements_df[versements_df['idMatricule'] == matricule]
        stats["total_verse"] = student_versements['Montant'].sum() if not student_versements.empty else 0
        stats["nombre_versements"] = len(student_versements)
    
    # Calcul des présences
    if not presences_df.empty and 'Matricule' in presences_df.columns:
        student_presences = presences_df[presences_df['Matricule'] == matricule]
        stats["total_seances"] = len(student_presences)
        
        if not student_presences.empty:
            stats["seances_presentes"] = len(student_presences[student_presences['Statut'] == 'Présent'])
            stats["seances_absentes"] = len(student_presences[student_presences['Statut'] == 'Absent'])
            stats["seances_retard"] = len(student_presences[student_presences['Statut'] == 'Retard'])
            
            if stats["total_seances"] > 0:
                stats["taux_absenteisme"] = (stats["seances_absentes"] / stats["total_seances"]) * 100
    
    return stats

def generate_receipt_html(student_info, versement_info):
    
    logo_sign = get_base64_image("Signature.png")
    logo_4 = get_base64_image("logo4.png")
    """Génère le HTML pour le reçu"""
    receipt_html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>Reçu de Paiement - LE BORELIEN</title>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 0; padding: 20px; background-color: #f5f5f5; }}
            .receipt {{ max-width: 600px; margin: 0 auto; background: white; padding: 30px; border-radius: 10px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); }}
            .header {{ text-align: center; border-bottom: 2px solid #1e3c72; padding-bottom: 20px; margin-bottom: 30px; }}
            .title {{ color: #1e3c72; font-size: 24px; font-weight: bold; margin-bottom: 10px; }}
            .subtitle {{ color: #666; font-size: 14px; }}
            .content {{ margin: 20px 0; }}
            .row {{ display: flex; justify-content: space-between; margin: 10px 0; padding: 8px 0; border-bottom: 1px solid #eee; }}
            .label {{ font-weight: bold; color: #333; }}
            .value {{ color: #666; }}
            .amount {{ font-size: 20px; font-weight: bold; color: #28a745; text-align: center; background: #f8f9fa; padding: 15px; border-radius: 5px; margin: 20px 0; }}
            .footer {{ text-align: center; margin-top: 30px; padding-top: 20px; border-top: 2px solid #1e3c72; font-size: 12px; color: #666; }}
            .stamp {{ text-align: right; margin-top: 40px; color: #1e3c72; font-weight: bold; }}
        </style>
    </head>
    <body>
        <div class="receipt">
            <div class="header">
                <div class="title">
                    🎓 LE BORELIEN
                </div>
                <div class="subtitle">
                    Groupe de soutien aux étudiants, préparation aux concours Bacc + 2, soutien scolaire à domicile
                </div>
                
                
                <div class="logo-section">
                    <img src="data:image/png;base64,{logo_4}" alt="Logo" style="height: 70px; width: 150px">
                </div>
            </div>
            
            <h2 style="text-align: center; color: #1e3c72; margin-bottom: 30px;">📄 REÇU DE PAIEMENT</h2>
            
            <div class="content">
                <div class="row">
                    <span class="label">N° de Reçu:</span>
                    <span class="value">REC-{datetime.now().strftime('%Y%m%d')}-{str(uuid.uuid4())[:8].upper()}</span>
                </div>
                <div class="row">
                    <span class="label">Date:</span>
                    <span class="value">{versement_info['date']}</span>
                </div>
                <div class="row">
                    <span class="label">Étudiant:</span>
                    <span class="value">{student_info['Nom']} {student_info['Prénom']}</span>
                </div>
                <div class="row">
                    <span class="label">Matricule:</span>
                    <span class="value">{student_info['Matricule']}</span>
                </div>
                <div class="row">
                    <span class="label">Filière:</span>
                    <span class="value">{student_info['Filière']}</span>
                </div>
                <div class="row">
                    <span class="label">Motif:</span>
                    <span class="value">{versement_info['motif'] if versement_info['motif'] else 'Frais de formation'}</span>
                </div>
            </div>
            
            <div class="amount">
                MONTANT VERSÉ: {versement_info['montant']:,.0f} FCFA
            </div>
            
            <div class="footer">
                <p><strong>Merci pour votre confiance !</strong></p>
                <p>📞 Contact: +237 6 97 20 32 87 // 6 76 58 51 83 | 📧 Email: leborelien@gmail.com</p>
                <p>🏢 Centres: Lycée de NGOA-EKELE</p>
            </div>
            
            <div class="stamp">
                <p>LE BORELIEN</p>
                <img src="data:image/png;base64,{logo_sign}" alt="Logo" style="height: 50px; width: 90px">
            </div>
        </div>
    </body>
    </html>
    """
    return receipt_html

def get_absent_students(start_date, end_date, min_absences, presences_df, etudiants_df):
    """Récupère la liste des étudiants les plus absentéistes sur une période"""
    if presences_df.empty or etudiants_df.empty:
        return pd.DataFrame()
    
    # Filtrer les présences sur la période
    presences_df['Date'] = pd.to_datetime(presences_df['Date'])
    filtered_presences = presences_df[
        (presences_df['Date'] >= pd.to_datetime(start_date)) &
        (presences_df['Date'] <= pd.to_datetime(end_date)) &
        (presences_df['Statut'] == 'Absent')
    ]
    
    if filtered_presences.empty:
        return pd.DataFrame()
    
    # Compter les absences par étudiant
    absence_counts = filtered_presences.groupby('Matricule').size().reset_index(name='NombreAbsences')
    
    # Filtrer par le nombre minimum d'absences
    absent_students = absence_counts[absence_counts['NombreAbsences'] >= min_absences]
    
    if absent_students.empty:
        return pd.DataFrame()
    
    # Joindre avec les informations des étudiants
    result = absent_students.merge(etudiants_df, on='Matricule', how='left')
    
    # Sélectionner et organiser les colonnes
    columns_to_show = ['Nom', 'Prénom', 'Filière', 'Téléphone', 'NombreAbsences']
    result = result[columns_to_show]
    result = result.sort_values('NombreAbsences', ascending=False)
    
    return result

def calculate_worked_hours(seances_df, teacher_ID, start_date, end_date):
    """Calcule les heures travaillées par un enseignant sur une période"""
    if seances_df.empty:
        return 0
    
    # Filtrer les séances de l'enseignant sur la période
    teacher_seances = seances_df[
        (seances_df['idEnseignant'] == teacher_ID) &
        (pd.to_datetime(seances_df['Date']) >= pd.to_datetime(start_date)) &
        (pd.to_datetime(seances_df['Date']) <= pd.to_datetime(end_date))
    ]
    
    total_hours = 0
    for _, seance in teacher_seances.iterrows():
        try:
            heure_a = datetime.strptime(str(seance['HeureArrivée']), '%H:%M').time()
            heure_d = datetime.strptime(str(seance['HeureDepart']), '%H:%M').time()
            duree_minutes = (datetime.combine(date.today(), heure_d) - 
                           datetime.combine(date.today(), heure_a)).seconds // 60
            total_hours += duree_minutes / 60
        except:
            # Durée par défaut si parsing échoue
            total_hours += 2
    
    return round(total_hours, 2)


def set_language():
    return st.sidebar.selectbox("🌍 Choisissez la langue / Choose the language", ["", "Français", "English"])

def generate_matricule():
    """Génère un matricule unique pour un étudiant"""
    return f"BRL25{str(uuid.uuid4())[:4].upper()}"

def show_pdf(path):
    with open(path,"rb") as f:
        base64_pdf=base64.b64encode(f.read()).decode('utf-8')
        pdf_display=f'<iframe src="data:application/pdf;base64, {base64_pdf}" width="700" height="500" type="application/pdf"></iframe>'
        st.markdown(pdf_display,unsafe_allow_html=True)
        

def generer_bulletin_paie_html(
    nom_employe,
    periode_debut,
    periode_fin,
    seances,
    montant_a_payer,
    acompte=0,
    prime=0,
    montant_percu=None
):
    """
    Génère un bulletin de paie en HTML
    
    Args:
        nom_employe (str): Nom de l'employé
        periode_debut (str): Date de début de période (format DD/MM/YYYY)
        periode_fin (str): Date de fin de période (format DD/MM/YYYY)
        seances (list): Liste des séances [{"date": "02/02/2025", "classe": "ISE ECO", "intitule": "Les indicateurs économiques"}]
        montant_a_payer (int): Montant total à payer
        acompte (int): Montant de l'acompte
        prime (int): Montant de la prime
        montant_percu (int): Montant déjà perçu (optionnel)
    
    Returns:
        str: Code HTML du bulletin de paie
    """
    
    nombre_seances = len(seances)
    montant_net = montant_a_payer + prime - acompte
    
    if montant_percu is None:
        montant_percu = montant_net
    
    # Génération des lignes du tableau des séances
    lignes_seances = ""
    for seance in seances:
        lignes_seances += f"""
        <tr>
            <td>{seance['date']}</td>
            <td>{seance['classe']}</td>
            <td>{seance['intitule']}</td>
        </tr>
        """
    
    # Génération des lignes vides pour remplir le tableau
    #lignes_vides = ""
    #for i in range(max(0, 10 - nombre_seances)):  # Ajouter des lignes vides pour un tableau uniforme
    #    lignes_vides += """
    #    <tr>
    #        <td>&nbsp;</td>
    #        <td>&nbsp;</td>
    #        <td>&nbsp;</td>
    #        <td>&nbsp;</td>
    #    </tr>
    #    """
    
    

# Encoder votre image
    logo_base64 = get_base64_image("logo3.png")
    logo_ruban = get_base64_image("logo_ruban.jpg")
    
    html_content = f"""
    <!DOCTYPE html>
    <html lang="fr">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Bulletin de Paie - {nom_employe}</title>
        <style>
            body {{
                font-family: Arial, sans-serif;
                margin: 20px;
                background-color: #f5f5f5;
            }}
            
            .bulletin-container {{
                max-width: 800px;
                margin: 0 auto;
                background-color: white;
                border-radius: 10px;
                overflow: hidden;
                box-shadow: 0 4px 8px rgba(0,0,0,0.1);
            }}
            
            .header {{
                background: linear-gradient(135deg, #4169E1, #1E90FF);
                color: white;
                padding: 20px;
                text-align: center;
                position: relative;
            }}
            
            .header h1 {{
                margin: 0;
                font-size: 28px;
                font-weight: bold;
            }}
            
            .logo-section {{
                position: absolute;
                right: 20px;
                top: 50%;
                transform: translateY(-50%);
                background-color: white;
                padding: 10px;
                border-radius: 8px;
                color: #e74c3c;
                font-weight: bold;
                font-size: 14px;
            }}
            
            .info-header {{
                background-color: #f8f9fa;
                padding: 15px 20px;
                border-bottom: 1px solid #dee2e6;
            }}
            
            .periode-info {{
                display: flex;
                justify-content: space-between;
                align-items: center;
                margin-bottom: 10px;
            }}
            
            .contact-info {{
                text-align: right;
                font-size: 14px;
                color: #6c757d;
            }}
            
            .employe-nom {{
                font-size: 24px;
                font-weight: bold;
                margin: 10px 0;
                color: #2c3e50;
            }}
            
            .section-title {{
                background-color: #2c3e50;
                color: white;
                padding: 12px;
                margin: 0;
                text-align: center;
                font-size: 18px;
                font-weight: bold;
            }}
            
            .seances-table {{
                width: 100%;
                border-collapse: collapse;
                margin-bottom: 20px;
            }}
            
            .seances-table th {{
                background-color: #e74c3c;
                color: white;
                padding: 12px;
                text-align: center;
                font-weight: bold;
            }}
            
            .seances-table td {{
                padding: 10px;
                border: 1px solid #dee2e6;
                text-align: left;
            }}
            
            .seances-table tr:nth-child(even) {{
                background-color: #f8f9fa;
            }}
            
            .calcul-section {{
                padding: 20px;
                background-color: #f8f9fa;
            }}
            
            .calcul-row {{
                display: flex;
                justify-content: space-between;
                align-items: center;
                margin: 10px 0;
                padding: 8px 0;
            }}
            
            .calcul-label {{
                font-weight: bold;
                color: #2c3e50;
            }}
            
            .calcul-value {{
                font-weight: bold;
                padding: 8px 15px;
                border-radius: 5px;
                min-width: 120px;
                text-align: center;
            }}
            
            .montant-payer {{ background-color: #fff3cd; border: 1px solid #ffeaa7; }}
            .acompte {{ background-color: #f8d7da; border: 1px solid #f5c6cb; }}
            .prime {{ background-color: #e2e3e5; border: 1px solid #d6d8db; }}
            .montant-net {{ background-color: #d4edda; border: 1px solid #c3e6cb; color: #155724; }}
            .montant-percu {{ background-color: #cce5ff; border: 1px solid #99d6ff; color: #0056b3; }}
            
            .note {{
                font-style: italic;
                color: #6c757d;
                font-size: 12px;
                margin: 15px 0;
                padding: 10px;
                background-color: #f8f9fa;
                border-left: 4px solid #6c757d;
            }}
            
            .footer {{
                background: linear-gradient(135deg, #2f13d8, #ee5a52);
                color: white;
                padding: 20px;
                text-align: center;
                font-weight: bold;
            }}
        </style>
    </head>
    <body>
        <div class="bulletin-container">
            <!-- En-tête -->
            <div class="header">
                <h1>Bulletin de paie</h1>
                <div class="logo-section">
                    <img src="data:image/png;base64,{logo_base64}" alt="Logo" style="height: 40px; width: 150px">
                </div>
            </div>
            
            <!-- Informations générales -->
            <div class="info-header">
                <div class="periode-info">
                    <div>
                        <strong>Période du :</strong> {periode_debut} <strong>Au :</strong> {periode_fin}
                    </div>
                    <div class="contact-info">
                        📧:leborelien@gmail.com<br>
                        Yaoundé, Lycée de NGOA-EKELE<br>
                        +237 6 97 20 32 87 // 6 76 58 51 83 |  
                    </div>
                </div>
                <div class="employe-nom">Mr. {nom_employe}</div>
            </div>
            
            <!-- Détails sur les séances -->
            <div class="section-title">Détails sur les séances</div>
            <table class="seances-table">
                <thead>
                    <tr>
                        <th>Date</th>
                        <th>Classe</th>
                        <th>Intitulé de la séance</th>
                    </tr>
                </thead>
                <tbody>
                    {lignes_seances}
                </tbody>
            </table>
            
            <!-- Calcul Financier -->
            <div class="section-title">Calcul Financier</div>
            <div class="calcul-section">
                <div class="calcul-row">
                    <span class="calcul-label">Nombre Total de Séance:</span>
                    <span class="calcul-value" style="background-color: white; border: 1px solid #dee2e6;">{nombre_seances}</span>
                </div>
                
                <div class="calcul-row">
                    <span class="calcul-label">Montant à payer :</span>
                    <span class="calcul-value montant-payer">{montant_a_payer:,} FCFA</span>
                </div>
                
                <div class="calcul-row">
                    <span class="calcul-label">Acompte :</span>
                    <span class="calcul-value acompte">{acompte} FCFA</span>
                </div>
                
                <div class="calcul-row">
                    <span class="calcul-label">Prime :</span>
                    <span class="calcul-value prime">{prime} FCFA</span>
                </div>
                
                <div class="note">
                    <strong>NB:</strong> Pour des soucis de monnaie, 
                    la différence entre la somme dans 
                    votre enveloppe et le montant net à 
                    percevoir sera déduit de votre 
                    prochaine paye.
                </div>
                
                <div class="calcul-row">
                    <span class="calcul-label">Montant Net à percevoir :</span>
                    <span class="calcul-value montant-net">{montant_net:,} FCFA</span>
                </div>
                
                <div class="calcul-row">
                    <span class="calcul-label">Montant perçu :</span>
                    <span class="calcul-value montant-percu">{montant_percu:,} FCFA</span>
                </div>
            </div>
            
            <!-- Pied de page -->
            <div class="footer">
                <img src="data:image/png;base64,{logo_ruban}" alt="Logo" style="height: 90px; width: 700px">
            </div>
        </div>
    </body>
    </html>
    """
    
    return html_content


# Fonction utilitaire pour sauvegarder le bulletin en fichier HTML
def sauvegarder_bulletin(html_content, nom_fichier):
    """
    Sauvegarde le bulletin de paie en fichier HTML
    
    Args:
        html_content (str): Contenu HTML du bulletin
        nom_fichier (str): Nom du fichier de sortie
    """
    with open(nom_fichier, 'w', encoding='utf-8') as f:
        f.write(html_content)
    print(f"Bulletin de paie sauvegardé dans {nom_fichier}")


def get_teacher_sessions(seances_df, teacher_ID, start_date, end_date):
    """
    Récupère les séances d'un enseignant pour une période donnée
    et les formate pour le bulletin de paie
    """
    # Filtrer les séances par enseignant et période
    teacher_sessions = seances_df[
        (seances_df['idEnseignant'] == teacher_ID) &
        (pd.to_datetime(seances_df['Date']).dt.date >= start_date) &
        (pd.to_datetime(seances_df['Date']).dt.date <= end_date)
    ].copy()
    
    # Formater pour le bulletin HTML
    sessions_formatted = []
    for _, session in teacher_sessions.iterrows():
        date_formatted = pd.to_datetime(session['Date']).strftime('%d/%m/%Y')
        sessions_formatted.append({
            "date": date_formatted,
            "classe": session.get('Classe', 'N/A'),
            "intitule": session.get('Sujet', session.get('IntituléCours', 'Cours'))
        })
    
    return sessions_formatted

       
def update_value_in_google_sheet(sheet_name, column_to_update, identifier_value, new_value, identifier_column=None):
    """
    Modifie une valeur spécifique dans une feuille Google Sheet
    
    Args:
        sheet_name (str): Nom de la feuille
        column_to_update (str): Nom de la colonne à modifier
        identifier_value: Valeur qui identifie la ligne à modifier
        new_value: Nouvelle valeur à insérer
        identifier_column (str, optional): Nom de la colonne d'identification. 
                                         Si None, utilise la première colonne
    
    Returns:
        bool: True si la modification a réussi, False sinon
    """
    try:
        worksheet = get_worksheet(sheet_name)
        if not worksheet:
            st.error(f"Impossible de trouver la feuille '{sheet_name}'")
            return False
        
        # Récupérer toutes les valeurs pour trouver la ligne et colonne
        values = worksheet.get_all_values()
        if not values or len(values) < 2:
            st.error(f"La feuille '{sheet_name}' est vide ou ne contient pas d'en-têtes")
            return False
        
        headers = values[0]
        
        # Déterminer la colonne d'identification (première colonne par défaut)
        if identifier_column is None:
            identifier_col_index = 0
            identifier_column = headers[0]
        else:
            try:
                identifier_col_index = headers.index(identifier_column)
            except ValueError:
                st.error(f"Colonne d'identification '{identifier_column}' non trouvée")
                return False
        
        # Trouver l'index de la colonne à modifier
        try:
            update_col_index = headers.index(column_to_update)
        except ValueError:
            st.error(f"Colonne '{column_to_update}' non trouvée")
            return False
        
        # Chercher la ligne correspondante
        target_row_index = None
        for i, row in enumerate(values[1:], start=2):  # start=2 car les en-têtes sont en ligne 1
            if len(row) > identifier_col_index and row[identifier_col_index] == str(identifier_value):
                target_row_index = i
                break
        
        if target_row_index is None:
            st.error(f"Aucune ligne trouvée avec {identifier_column} = '{identifier_value}'")
            return False
        
        # Calculer l'adresse de la cellule (ex: B5)
        # Convertir l'index de colonne en lettre (A=0, B=1, etc.)
        col_letter = chr(ord('A') + update_col_index)
        cell_address = f"{col_letter}{target_row_index}"
        
        # Mettre à jour la cellule
        worksheet.update(cell_address, new_value)
        
        st.success(f"Valeur mise à jour avec succès dans {sheet_name}[{cell_address}]: '{new_value}'")
        return True
        
    except Exception as e:
        st.error(f"Erreur lors de la modification dans '{sheet_name}': {e}")
        return False

 
def update_value_in_google_sheet2(sheet_name, column_to_update, identifier_value, new_value, identifier_column=None):
    """
    Modifie une valeur spécifique dans une feuille Google Sheet
    
    Args:
        sheet_name (str): Nom de la feuille
        column_to_update (str): Nom de la colonne à modifier
        identifier_value: Valeur qui identifie la ligne à modifier
        new_value: Nouvelle valeur à insérer
        identifier_column (str, optional): Nom de la colonne d'identification. 
                                         Si None, utilise la première colonne
    
    Returns:
        bool: True si la modification a réussi, False sinon
    """
    try:
        worksheet = get_worksheet(sheet_name)
        if not worksheet:
            st.error(f"Impossible de trouver la feuille '{sheet_name}'")
            return False
        
        # Récupérer toutes les valeurs pour trouver la ligne et colonne
        values = worksheet.get_all_values()
        if not values or len(values) < 2:
            st.error(f"La feuille '{sheet_name}' est vide ou ne contient pas d'en-têtes")
            return False
        
        headers = values[0]
        
        # Déterminer la colonne d'identification (première colonne par défaut)
        if identifier_column is None:
            identifier_col_index = 0
            identifier_column = headers[0]
        else:
            try:
                identifier_col_index = headers.index(identifier_column)
            except ValueError:
                st.error(f"Colonne d'identification '{identifier_column}' non trouvée")
                return False
        
        # Trouver l'index de la colonne à modifier
        try:
            update_col_index = headers.index(column_to_update)
        except ValueError:
            st.error(f"Colonne '{column_to_update}' non trouvée")
            return False
        
        # Chercher la ligne correspondante
        target_row_index = None
        for i, row in enumerate(values[1:], start=2):  # start=2 car les en-têtes sont en ligne 1
            if len(row) > identifier_col_index and row[identifier_col_index] == str(identifier_value):
                target_row_index = i
                break
        
        if target_row_index is None:
            st.error(f"Aucune ligne trouvée avec {identifier_column} = '{identifier_value}'")
            return False
        
        # Calculer l'adresse de la cellule (ex: B5)
        # Convertir l'index de colonne en lettre (A=0, B=1, etc.)
        col_letter = chr(ord('A') + update_col_index)
        cell_address = f"{col_letter}{target_row_index}"
        
        # Mettre à jour la cellule - conversion en string pour éviter les erreurs de format
        worksheet.update(cell_address, [[str(new_value)]])
        
        st.success(f"Valeur mise à jour avec succès dans {sheet_name}[{cell_address}]: '{new_value}'")
        return True
        
    except Exception as e:
        st.error(f"Erreur lors de la modification dans '{sheet_name}': {e}")
        return False 
       
